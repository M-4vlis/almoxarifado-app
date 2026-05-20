import argparse
import math
import os
import re
import time
import unicodedata
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


# =========================
# CONFIGURACOES
# =========================

ARQUIVO_MATERIAIS = "data/LISTA de MATERIAIS.xlsx"
ARQUIVO_COLABORADORES = "data/DADOS COLABORADORES.xlsx"
ARQUIVO_GLPI_REQUISICAO = "data/GLPI - REQUISICAO.xlsx"

ABAS_MATERIAIS = [
    "CENTRAL",
    "SUB",
    "BENFICA",
]

LIMITE_BATCH = 450
LIMITE_BATCH_SOLICITACOES = 300
PAUSA_ENTRE_LOTES_SOLICITACOES = 0.4

firebase_admin = None
auth = None
credentials = None
firestore = None
FieldFilter = None
ResourceExhausted = None


# =========================
# FUNCOES AUXILIARES
# =========================

def limpar_numero(valor):
    if pd.isna(valor):
        return ""

    texto = str(valor).strip()

    if texto.endswith(".0"):
        texto = texto[:-2]

    return re.sub(r"\D", "", texto)


def limpar_texto(valor):
    if pd.isna(valor):
        return ""

    return str(valor).strip()


def normalizar_cabecalho(valor):
    texto = limpar_texto(valor).lower()
    texto = unicodedata.normalize("NFKD", texto)
    texto = "".join(
        caractere
        for caractere in texto
        if not unicodedata.combining(caractere)
    )
    texto = re.sub(r"\s+", " ", texto)

    return texto.strip()


def obter_coluna(colunas, nomes, obrigatoria=True):
    mapa_colunas = {
        normalizar_cabecalho(coluna): coluna
        for coluna in colunas
    }

    for nome in nomes:
        coluna = mapa_colunas.get(normalizar_cabecalho(nome))

        if coluna is not None:
            return coluna

    if obrigatoria:
        raise Exception(
            f"Coluna obrigatoria nao encontrada: {nomes[0]}"
        )

    return None


def normalizar_perfil(valor):
    perfil = limpar_texto(valor).lower()

    if perfil in ["admin", "usuario"]:
        return perfil

    return "usuario"


def normalizar_numero_monetario(valor):
    if valor is None:
        return 0

    try:
        if pd.isna(valor):
            return 0
    except (TypeError, ValueError):
        pass

    if isinstance(valor, (int, float)):
        numero = float(valor)

        if not math.isfinite(numero):
            return 0

        return numero

    if isinstance(valor, str):
        texto = valor.strip()

        if texto.lower() in ["", "-", "nan", "none", "null"]:
            return 0

        texto = texto.replace("R$", "")
        texto = texto.replace("r$", "")
        texto = texto.replace("\xa0", "")
        texto = texto.replace(" ", "")
        texto = re.sub(r"[^0-9,.\-]", "", texto)

        if texto in ["", "-", ".", ","]:
            return 0

        if "," in texto and "." in texto:
            if texto.rfind(",") > texto.rfind("."):
                texto = texto.replace(".", "")
                texto = texto.replace(",", ".")
            else:
                texto = texto.replace(",", "")

        elif "," in texto:
            texto = texto.replace(".", "")
            texto = texto.replace(",", ".")

        elif texto.count(".") > 1:
            texto = texto.replace(".", "")

        elif "." in texto:
            partes = texto.split(".")

            if len(partes[-1]) == 3 and len(partes[0]) <= 3:
                texto = texto.replace(".", "")

        valor = texto

    try:
        numero = float(valor)
    except (TypeError, ValueError):
        return 0

    if not math.isfinite(numero):
        return 0

    return numero


def converter_valor_brasileiro(valor):
    return normalizar_numero_monetario(valor)


def formatar_valor_unitario(valor):
    numero = normalizar_numero_monetario(valor)

    texto = f"R$ {numero:,.2f}"
    texto = texto.replace(",", "X")
    texto = texto.replace(".", ",")
    texto = texto.replace("X", ".")

    return texto


def preparar_valor_unitario(valor):
    numero = normalizar_numero_monetario(valor)

    return {
        "valorUnitario": formatar_valor_unitario(numero),
        "valorUnitarioNumero": numero,
    }


def formatar_valor_excel_com_zeros(celula):
    valor = celula.value

    if valor is None:
        return ""

    formato = str(celula.number_format)

    if isinstance(valor, str):
        return valor.strip()

    if isinstance(valor, float) and valor.is_integer():
        valor = int(valor)

    texto = str(valor).strip()

    if texto.endswith(".0"):
        texto = texto[:-2]

    texto = re.sub(r"\D", "", texto)

    quantidade_zeros_formato = formato.count("0")

    if quantidade_zeros_formato > len(texto):
        texto = texto.zfill(quantidade_zeros_formato)

    return texto


def normalizar_id(valor):
    texto = str(valor or "").strip().upper()
    texto = re.sub(r"[^A-Z0-9_-]", "_", texto)
    texto = re.sub(r"_+", "_", texto)

    return texto.strip("_")


def montar_email_login(matricula):
    return f"{matricula}@almox.local"


def carregar_env_local():
    caminho_env = Path(".env")

    if not caminho_env.exists():
        return

    for linha in caminho_env.read_text(encoding="utf-8").splitlines():
        linha = linha.strip()

        if not linha or linha.startswith("#") or "=" not in linha:
            continue

        chave, valor = linha.split("=", 1)
        chave = chave.strip()
        valor = valor.strip().strip('"').strip("'")

        if chave and chave not in os.environ:
            os.environ[chave] = valor


def obter_lista_env(nome):
    valor = os.getenv(nome, "")

    if not valor:
        return []

    partes = re.split(r"[\s,;]+", valor)

    return [
        parte.strip()
        for parte in partes
        if parte.strip()
    ]


def carregar_firebase_admin():
    global firebase_admin
    global auth
    global credentials
    global firestore
    global FieldFilter
    global ResourceExhausted

    if firebase_admin:
        return

    try:
        import firebase_admin as firebase_admin_modulo
        from firebase_admin import auth as auth_modulo
        from firebase_admin import credentials as credentials_modulo
        from firebase_admin import firestore as firestore_modulo
        from google.cloud.firestore_v1.base_query import FieldFilter as FieldFilterModulo
        from google.api_core.exceptions import ResourceExhausted as ResourceExhaustedModulo
    except ModuleNotFoundError as erro:
        raise ModuleNotFoundError(
            "Dependencia firebase-admin nao instalada. Rode: pip install -r requirements.txt"
        ) from erro

    firebase_admin = firebase_admin_modulo
    auth = auth_modulo
    credentials = credentials_modulo
    firestore = firestore_modulo
    FieldFilter = FieldFilterModulo
    ResourceExhausted = ResourceExhaustedModulo


def inicializar_firebase(caminho_service_account):
    carregar_firebase_admin()

    if firebase_admin._apps:
        return

    if caminho_service_account:
        credencial = credentials.Certificate(caminho_service_account)
        firebase_admin.initialize_app(credencial)
        return

    firebase_admin.initialize_app()


def confirmar_arquivo(caminho):
    arquivo = Path(caminho)

    if not arquivo.exists():
        raise FileNotFoundError(f"Arquivo nao encontrado: {caminho}")

    return arquivo


def commitar_batch(batch, contador, dry_run):
    if contador == 0:
        return 0

    if dry_run:
        return 0

    batch.commit()

    return 0


def set_em_lotes(db, colecao, documentos, dry_run):
    if dry_run:
        return len(documentos)

    batch = db.batch()
    contador = 0
    total = 0

    for doc_id, dados in documentos:
        total += 1

        if not dry_run:
            referencia = db.collection(colecao).document(doc_id)
            batch.set(
                referencia,
                dados,
                merge=True,
            )

        contador += 1

        if contador >= LIMITE_BATCH:
            contador = commitar_batch(
                batch,
                contador,
                dry_run,
            )
            batch = db.batch()

    commitar_batch(
        batch,
        contador,
        dry_run,
    )

    return total


def timestamp_servidor(dry_run):
    if dry_run:
        return "DRY_RUN_SERVER_TIMESTAMP"

    return firestore.SERVER_TIMESTAMP


def erro_quota_firestore(erro):
    if ResourceExhausted and isinstance(erro, ResourceExhausted):
        return True

    texto = str(erro).lower()

    return (
        "429" in texto or
        "quota" in texto or
        "resourceexhausted" in texto or
        "resource exhausted" in texto
    )


# =========================
# LER PLANILHAS
# =========================

def ler_materiais():
    confirmar_arquivo(ARQUIVO_MATERIAIS)

    materiais_dict = {}

    for aba in ABAS_MATERIAIS:
        df = pd.read_excel(
            ARQUIVO_MATERIAIS,
            sheet_name=aba,
        )

        coluna_codigo = obter_coluna(
            df.columns,
            [
                "Cod",
                "CÃ³d",
                "Codigo",
                "Código",
            ],
        )

        coluna_material = obter_coluna(
            df.columns,
            [
                "Material",
            ],
        )

        coluna_estoque = obter_coluna(
            df.columns,
            [
                "Estoque",
            ],
        )

        coluna_valor_unitario = obter_coluna(
            df.columns,
            [
                "Valor Unitario",
                "Valor unitario",
                "VALOR UNITARIO",
                "Valor Unitário",
                "Valor unitário",
                "VALOR UNITÁRIO",
                "Valor Unitário",
                "Valor unitário",
                "VALOR UNITÁRIO",
                "Valor Unitario",
                "Valor",
                "VALOR",
                "Valor Unitário",
                "Valor UnitÃ¡rio",
            ],
            obrigatoria=False,
        )

        if not coluna_valor_unitario:
            print(
                "AVISO: Coluna de valor unitario nao encontrada "
                f"na aba {aba}. Importando valores como 0."
            )

        else:
            print(
                f"Aba {aba}: coluna de valor unitario encontrada: "
                f"{coluna_valor_unitario}"
            )

        df = df[
            [
                coluna_codigo,
                coluna_material,
                coluna_estoque,
                *(
                    [coluna_valor_unitario]
                    if coluna_valor_unitario
                    else []
                ),
            ]
        ]

        df = df.dropna(
            subset=[
                coluna_codigo,
                coluna_material,
            ]
        )

        for _, linha in df.iterrows():
            codigo = limpar_numero(linha[coluna_codigo])
            descricao = limpar_texto(linha[coluna_material])

            if not codigo or not descricao:
                continue

            estoque = linha[coluna_estoque]

            if pd.isna(estoque):
                estoque = 0

            try:
                estoque = float(estoque)
            except (TypeError, ValueError):
                estoque = 0

            valor_unitario = preparar_valor_unitario(0)

            if coluna_valor_unitario:
                valor_unitario = preparar_valor_unitario(
                    linha[coluna_valor_unitario]
                )

            doc_id = f"{codigo}_{normalizar_id(aba)}"

            material_existente = materiais_dict.get(doc_id)

            dados = {
                "codigo": codigo,
                "descricao": descricao,
                "almoxarifado": aba,
                "estoque": estoque,
                "disponivel": estoque > 0,
                "valorUnitario": valor_unitario["valorUnitario"],
                "valorUnitarioNumero": valor_unitario["valorUnitarioNumero"],
                "ativo": True,
                "origem": "planilha_materiais",
            }

            if not material_existente:
                materiais_dict[doc_id] = dados
                continue

            if estoque > float(material_existente.get("estoque") or 0):
                material_existente["estoque"] = estoque

            if estoque > 0:
                material_existente["disponivel"] = True

            if len(descricao) > len(material_existente.get("descricao") or ""):
                material_existente["descricao"] = descricao

            valor_existente = normalizar_numero_monetario(
                material_existente.get("valorUnitarioNumero") or
                material_existente.get("valorUnitario")
            )

            if (
                valor_unitario["valorUnitarioNumero"] > 0 and
                valor_unitario["valorUnitarioNumero"] > valor_existente
            ):
                material_existente["valorUnitario"] = valor_unitario["valorUnitario"]
                material_existente["valorUnitarioNumero"] = valor_unitario["valorUnitarioNumero"]

    return materiais_dict


def ler_colaboradores():
    caminho = confirmar_arquivo(ARQUIVO_COLABORADORES)

    workbook = load_workbook(
        caminho,
        data_only=True,
    )

    planilha = workbook.active
    cabecalhos = {}

    for coluna in range(1, planilha.max_column + 1):
        valor = planilha.cell(
            row=1,
            column=coluna,
        ).value

        if valor is not None:
            cabecalhos[str(valor).strip()] = coluna

    coluna_nome = obter_coluna(
        cabecalhos.keys(),
        [
            "NOME",
        ],
    )

    coluna_matricula = obter_coluna(
        cabecalhos.keys(),
        [
            "MATRÍCULA",
            "MATRICULA",
        ],
    )

    coluna_cpf = obter_coluna(
        cabecalhos.keys(),
        [
            "CPF",
        ],
    )

    coluna_perfil = obter_coluna(
        cabecalhos.keys(),
        [
            "PERFIL",
        ],
        obrigatoria=False,
    )

    colaboradores = []

    for linha in range(2, planilha.max_row + 1):
        nome = limpar_texto(
            planilha.cell(
                row=linha,
                column=cabecalhos[coluna_nome],
            ).value
        )

        matricula = formatar_valor_excel_com_zeros(
            planilha.cell(
                row=linha,
                column=cabecalhos[coluna_matricula],
            )
        )

        cpf = formatar_valor_excel_com_zeros(
            planilha.cell(
                row=linha,
                column=cabecalhos[coluna_cpf],
            )
        )

        perfil = "usuario"

        if coluna_perfil:
            perfil = normalizar_perfil(
                planilha.cell(
                    row=linha,
                    column=cabecalhos[coluna_perfil],
                ).value
            )

        if not nome or not matricula or not cpf:
            continue

        colaboradores.append(
            {
                "nome": nome,
                "matricula": matricula,
                "cpf": cpf,
                "emailLogin": montar_email_login(matricula),
                "perfil": perfil,
            }
        )

    return colaboradores


def ler_vinculos_requisicoes():
    confirmar_arquivo(ARQUIVO_GLPI_REQUISICAO)

    df = pd.read_excel(ARQUIVO_GLPI_REQUISICAO)

    colunas_necessarias = [
        "GLPI",
        "NÚMERO DA REQUISIÇÃO",
    ]

    for coluna in colunas_necessarias:
        if coluna not in df.columns:
            raise Exception(
                f"Coluna obrigatoria nao encontrada na planilha GLPI x Requisicao: {coluna}"
            )

    df = df.dropna(
        subset=[
            "GLPI",
            "NÚMERO DA REQUISIÇÃO",
        ]
    )

    mapa = {}

    for _, linha in df.iterrows():
        glpi = limpar_numero(linha["GLPI"])
        numero_requisicao = limpar_numero(linha["NÚMERO DA REQUISIÇÃO"])

        if not glpi or not numero_requisicao:
            continue

        if glpi not in mapa:
            mapa[glpi] = []

        if numero_requisicao not in mapa[glpi]:
            mapa[glpi].append(numero_requisicao)

    return mapa


# =========================
# IMPORTACOES FIREBASE
# =========================

def importar_materiais(db, dry_run, desativar_ausentes):
    materiais = ler_materiais()
    documentos = []

    for doc_id, material in materiais.items():
        documentos.append(
            (
                doc_id,
                {
                    **material,
                    "atualizadoEm": timestamp_servidor(dry_run),
                },
            )
        )

    total = set_em_lotes(
        db,
        "materiais",
        documentos,
        dry_run,
    )

    desativados = 0

    if desativar_ausentes and not dry_run:
        ids_atuais = set(materiais.keys())
        batch = db.batch()
        contador = 0

        for snapshot in db.collection("materiais").stream():
            if snapshot.id in ids_atuais:
                continue

            batch.set(
                snapshot.reference,
                {
                    "ativo": False,
                    "atualizadoEm": timestamp_servidor(dry_run),
                },
                merge=True,
            )

            contador += 1
            desativados += 1

            if contador >= LIMITE_BATCH:
                batch.commit()
                batch = db.batch()
                contador = 0

        if contador:
            batch.commit()

    print(f"Materiais lidos/importados: {total}")

    if desativar_ausentes:
        print(f"Materiais ausentes desativados: {desativados}")

    if not dry_run:
        db.collection("configuracoes").document("materiais").set(
            {
                "totalAtivos": len(materiais),
                "atualizadoEm": timestamp_servidor(dry_run),
                "origem": "importar_firebase",
            },
            merge=True,
        )


def obter_usuario_auth(email):
    try:
        return auth.get_user_by_email(email)
    except auth.UserNotFoundError:
        return None


def importar_usuarios(db, dry_run, salvar_cpf, atualizar_senhas, matriculas_admin):
    colaboradores = ler_colaboradores()
    criados = 0
    atualizados = 0
    matriculas_admin = set(matriculas_admin or [])

    for colaborador in colaboradores:
        email = colaborador["emailLogin"]
        senha = limpar_numero(colaborador["cpf"])
        usuario_auth = None

        if not dry_run:
            usuario_auth = obter_usuario_auth(email)

        if usuario_auth:
            atualizados += 1

            parametros = {
                "display_name": colaborador["nome"],
                "disabled": False,
            }

            if atualizar_senhas:
                parametros["password"] = senha

            auth.update_user(
                usuario_auth.uid,
                **parametros,
            )

            uid = usuario_auth.uid

        else:
            criados += 1

            if dry_run:
                uid = f"dry_run_{colaborador['matricula']}"
            else:
                novo_usuario = auth.create_user(
                    email=email,
                    password=senha,
                    display_name=colaborador["nome"],
                    disabled=False,
                )

                uid = novo_usuario.uid

        documento_existente = {}

        if not dry_run:
            snapshot = db.collection("usuarios").document(uid).get()

            if snapshot.exists:
                documento_existente = snapshot.to_dict() or {}

        perfil_padrao = colaborador.get("perfil", "usuario")

        if colaborador["matricula"] in matriculas_admin:
            perfil_padrao = "admin"

        dados_usuario = {
            "uid": uid,
            "nome": colaborador["nome"],
            "matricula": colaborador["matricula"],
            "emailLogin": email,
            "perfil": perfil_padrao,
            "ativo": documento_existente.get("ativo", True),
            "origem": "planilha_colaboradores",
            "atualizadoEm": timestamp_servidor(dry_run),
        }

        if not dry_run:
            referencia_usuario = db.collection("usuarios").document(uid)

            if salvar_cpf:
                dados_usuario["cpf"] = colaborador["cpf"]

            referencia_usuario.set(
                dados_usuario,
                merge=True,
            )

            if not salvar_cpf and "cpf" in documento_existente:
                referencia_usuario.update(
                    {
                        "cpf": firestore.DELETE_FIELD,
                    }
                )

    print(f"Usuarios criados: {criados}")
    print(f"Usuarios atualizados: {atualizados}")
    print(f"Total de colaboradores lidos: {len(colaboradores)}")


def importar_vinculos_requisicoes(db, dry_run):
    mapa = ler_vinculos_requisicoes()
    documentos_individuais = []
    documentos_por_glpi = []

    for glpi, requisicoes in mapa.items():
        requisicoes_ordenadas = sorted(requisicoes)

        documentos_por_glpi.append(
            (
                glpi,
                {
                    "glpi": glpi,
                    "numeroRequisicao": requisicoes_ordenadas[0],
                    "requisicoesVinculadas": requisicoes_ordenadas,
                    "totalRequisicoes": len(requisicoes_ordenadas),
                    "ativo": True,
                    "origem": "planilha_glpi_requisicao",
                    "atualizadoEm": timestamp_servidor(dry_run),
                },
            )
        )

        for numero_requisicao in requisicoes_ordenadas:
            doc_id = f"{glpi}_{numero_requisicao}"

            documentos_individuais.append(
                (
                    doc_id,
                    {
                        "glpi": glpi,
                        "numeroRequisicao": numero_requisicao,
                        "ativo": True,
                        "origem": "planilha_glpi_requisicao",
                        "atualizadoEm": timestamp_servidor(dry_run),
                    },
                )
            )

    total_individuais = set_em_lotes(
        db,
        "vinculosRequisicoes",
        documentos_individuais,
        dry_run,
    )

    total_por_glpi = set_em_lotes(
        db,
        "requisicoesPorGlpi",
        documentos_por_glpi,
        dry_run,
    )

    print(f"Vinculos individuais importados: {total_individuais}")
    print(f"GLPIs com requisicao importadas: {total_por_glpi}")

    return mapa


def solicitacao_precisa_atualizar(dados, requisicoes_ordenadas):
    requisicoes_atuais = dados.get("requisicoesVinculadas") or []

    if not isinstance(requisicoes_atuais, list):
        requisicoes_atuais = [requisicoes_atuais]

    requisicoes_atuais = [
        limpar_numero(requisicao)
        for requisicao in requisicoes_atuais
        if limpar_numero(requisicao)
    ]

    return not (
        limpar_numero(dados.get("numeroRequisicao")) == requisicoes_ordenadas[0] and
        requisicoes_atuais == requisicoes_ordenadas and
        dados.get("statusAtendimento") == "requisicao_vinculada" and
        dados.get("status") == "requisicao_vinculada"
    )


def imprimir_resumo_atualizacao_solicitacoes(resumo):
    print("\nResumo da atualizacao de solicitacoes")
    print(f"GLPIs lidas: {resumo['glpis_lidas']}")
    print(f"Vinculos importados: {resumo['vinculos_importados']}")
    print(f"Solicitacoes encontradas: {resumo['solicitacoes_encontradas']}")
    print(f"Solicitacoes atualizadas: {resumo['solicitacoes_atualizadas']}")
    print(f"Solicitacoes sem correspondencia: {resumo['solicitacoes_sem_correspondencia']}")

    if resumo["avisos_quota"]:
        print(f"Avisos de quota: {resumo['avisos_quota']}")


def atualizar_solicitacoes_com_requisicoes(db, mapa, dry_run):
    resumo = {
        "glpis_lidas": len(mapa),
        "vinculos_importados": sum(len(requisicoes) for requisicoes in mapa.values()),
        "solicitacoes_encontradas": 0,
        "solicitacoes_atualizadas": 0,
        "solicitacoes_sem_correspondencia": 0,
        "avisos_quota": 0,
    }

    if dry_run and db is None:
        print("Solicitacoes nao consultadas no dry-run sem Firebase.")
        imprimir_resumo_atualizacao_solicitacoes(resumo)
        return

    batch = None
    contador_batch = 0

    if not dry_run:
        batch = db.batch()

    def commitar_lote_solicitacoes():
        nonlocal batch
        nonlocal contador_batch

        if contador_batch == 0:
            return

        if dry_run:
            resumo["solicitacoes_atualizadas"] += contador_batch
            contador_batch = 0
            return

        try:
            batch.commit()
            resumo["solicitacoes_atualizadas"] += contador_batch
            time.sleep(PAUSA_ENTRE_LOTES_SOLICITACOES)
        except Exception as erro:
            if erro_quota_firestore(erro):
                print(
                    "\nAVISO: quota do Firestore excedida durante a gravacao "
                    "das solicitacoes. Aguarde alguns minutos e rode novamente, "
                    "ou reduza o volume da importacao."
                )
                raise

            raise

        batch = db.batch()
        contador_batch = 0

    for glpi, requisicoes in mapa.items():
        requisicoes_ordenadas = sorted(requisicoes)

        try:
            snapshots = db.collection("solicitacoes").where(
                "glpi",
                "==",
                glpi,
            ).stream()

            encontradas_glpi = 0

            for snapshot in snapshots:
                encontradas_glpi += 1
                resumo["solicitacoes_encontradas"] += 1

                dados = snapshot.to_dict() or {}

                if not solicitacao_precisa_atualizar(
                    dados,
                    requisicoes_ordenadas,
                ):
                    continue

                if not dry_run:
                    batch.set(
                        snapshot.reference,
                        {
                            "numeroRequisicao": requisicoes_ordenadas[0],
                            "requisicoesVinculadas": requisicoes_ordenadas,
                            "statusAtendimento": "requisicao_vinculada",
                            "status": "requisicao_vinculada",
                            "atualizadoEm": timestamp_servidor(dry_run),
                        },
                        merge=True,
                    )

                contador_batch += 1

                if contador_batch >= LIMITE_BATCH_SOLICITACOES:
                    commitar_lote_solicitacoes()

            if encontradas_glpi == 0:
                resumo["solicitacoes_sem_correspondencia"] += 1

            time.sleep(PAUSA_ENTRE_LOTES_SOLICITACOES)

        except Exception as erro:
            if erro_quota_firestore(erro):
                resumo["avisos_quota"] += 1
                print(
                    "\nAVISO: quota do Firestore excedida ao consultar "
                    f"solicitacoes da GLPI {glpi}. Aguarde alguns minutos "
                    "e rode novamente, ou reduza o volume da importacao."
                )
                break

            raise

    try:
        commitar_lote_solicitacoes()
    except Exception as erro:
        if erro_quota_firestore(erro):
            resumo["avisos_quota"] += 1
        else:
            raise

    imprimir_resumo_atualizacao_solicitacoes(resumo)


def obter_numero_seguro(valor):
    numero = normalizar_numero_monetario(valor)

    if not math.isfinite(numero):
        return 0

    return numero


def chave_material(codigo, almoxarifado):
    return (
        limpar_numero(codigo),
        limpar_texto(almoxarifado).upper(),
    )


def carregar_mapa_materiais_firestore(db):
    materiais_por_chave = {}

    for snapshot in db.collection("materiais").stream():
        dados = snapshot.to_dict() or {}
        chave = chave_material(
            dados.get("codigo"),
            dados.get("almoxarifado"),
        )

        if not chave[0] or not chave[1]:
            continue

        materiais_por_chave[chave] = obter_numero_seguro(
            dados.get("valorUnitarioNumero") or dados.get("valorUnitario")
        )

    return materiais_por_chave


def obter_valor_unitario_item(item, materiais_por_chave=None):
    valor = obter_numero_seguro(
        item.get("valorUnitarioNumero") or item.get("valorUnitario")
    )

    if valor > 0:
        return valor

    if materiais_por_chave:
        valor_material = obter_numero_seguro(
            materiais_por_chave.get(
                chave_material(
                    item.get("codigo"),
                    item.get("almoxarifado"),
                )
            )
        )

        if valor_material > 0:
            return valor_material

    quantidade = obter_numero_seguro(item.get("quantidade"))
    subtotal = obter_numero_seguro(
        item.get("subtotal") or item.get("valorTotalItem")
    )

    if quantidade > 0 and subtotal > 0:
        return subtotal / quantidade

    return 0


def calcular_subtotal_item(item, materiais_por_chave=None):
    quantidade = obter_numero_seguro(item.get("quantidade"))
    valor_unitario = obter_valor_unitario_item(item, materiais_por_chave)

    if quantidade > 0 and valor_unitario > 0:
        return quantidade * valor_unitario

    return obter_numero_seguro(
        item.get("subtotal") or item.get("valorTotalItem")
    )


def obter_total_solicitacao(dados, materiais_por_chave=None):
    total = obter_numero_seguro(dados.get("valorTotal"))

    if total > 0:
        return total

    total_itens = sum(
        calcular_subtotal_item(item, materiais_por_chave)
        for item in dados.get("itens") or []
        if isinstance(item, dict)
    )

    if total_itens > 0:
        return total_itens

    return obter_numero_seguro(
        dados.get("valorTotalEstimado") or dados.get("totalEstimado")
    )


def obter_quantidade_item(item):
    return obter_numero_seguro(item.get("quantidade"))


def solicitacao_tem_requisicao(dados):
    return bool(
        dados.get("numeroRequisicao") or
        dados.get("requisicoesVinculadas") or
        dados.get("statusAtendimento") in ["requisicao_vinculada", "concluida"]
    )


def somar_ranking(mapa, chave, quantidade=0, total=0, detalhe=""):
    chave_final = limpar_texto(chave) or "Nao informado"

    if chave_final not in mapa:
        mapa[chave_final] = {
            "label": chave_final,
            "detalhe": detalhe,
            "quantidade": 0,
            "total": 0,
        }

    mapa[chave_final]["quantidade"] += quantidade
    mapa[chave_final]["total"] += total

    if detalhe and not mapa[chave_final].get("detalhe"):
        mapa[chave_final]["detalhe"] = detalhe


def obter_top_ranking(mapa, campo, limite=5):
    return sorted(
        mapa.values(),
        key=lambda item: obter_numero_seguro(item.get(campo)),
        reverse=True,
    )[:limite]


def obter_data_ordenacao_solicitacao(dados):
    criado_em = dados.get("criadoEm")

    if hasattr(criado_em, "timestamp"):
        return criado_em.timestamp()

    return 0


def recalcular_resumo_admin(db, dry_run):
    if dry_run:
        print("Resumo admin nao recalculado no dry-run.")
        return

    total_solicitacoes = 0
    aguardando = 0
    vinculadas = 0
    total_itens = 0
    total_estimado = 0
    usuarios = set()
    glpis = set()
    requisicoes = set()
    usuarios_quantidade = {}
    usuarios_valor = {}
    glpis_valor = {}
    requisicoes_valor = {}
    materiais_quantidade = {}
    almoxarifados_quantidade = {}
    recentes = []
    materiais_por_chave = carregar_mapa_materiais_firestore(db)

    for snapshot in db.collection("solicitacoes").stream():
        dados = snapshot.to_dict() or {}
        total_solicitacoes += 1

        status_atendimento = dados.get("statusAtendimento") or ""

        if status_atendimento == "aguardando_requisicao":
            aguardando += 1

        if solicitacao_tem_requisicao(dados):
            vinculadas += 1

        total = obter_total_solicitacao(dados, materiais_por_chave)
        total_estimado += total

        nome_usuario = (
            dados.get("usuarioNome") or
            (dados.get("usuarioSolicitante") or {}).get("nome") or
            "Nao informado"
        )
        matricula_usuario = (
            dados.get("usuarioMatricula") or
            (dados.get("usuarioSolicitante") or {}).get("matricula") or
            ""
        )
        chave_usuario = (
            f"{nome_usuario} ({matricula_usuario})"
            if matricula_usuario
            else nome_usuario
        )

        if dados.get("usuarioUid") or matricula_usuario or nome_usuario:
            usuarios.add(dados.get("usuarioUid") or matricula_usuario or nome_usuario)

        glpi = limpar_texto(dados.get("glpi"))

        if glpi:
            glpis.add(glpi)

        lista_requisicoes = [
            dados.get("numeroRequisicao"),
            *(dados.get("requisicoesVinculadas") or []),
        ]

        lista_requisicoes = [
            limpar_texto(requisicao)
            for requisicao in lista_requisicoes
            if limpar_texto(requisicao)
        ]

        for requisicao in set(lista_requisicoes):
            requisicoes.add(requisicao)
            somar_ranking(
                requisicoes_valor,
                requisicao,
                quantidade=1,
                total=total,
                detalhe=f"GLPI {glpi or '-'}",
            )

        somar_ranking(
            usuarios_quantidade,
            chave_usuario,
            quantidade=1,
            total=total,
        )
        somar_ranking(
            usuarios_valor,
            chave_usuario,
            quantidade=1,
            total=total,
        )
        somar_ranking(
            glpis_valor,
            glpi or "Sem GLPI",
            quantidade=1,
            total=total,
            detalhe=f"{dados.get('totalItens') or 0} item(ns)",
        )

        for item in dados.get("itens") or []:
            if not isinstance(item, dict):
                continue

            quantidade = obter_quantidade_item(item)
            subtotal = calcular_subtotal_item(item, materiais_por_chave)
            total_itens += quantidade

            somar_ranking(
                materiais_quantidade,
                f"{item.get('codigo') or '-'} - {item.get('descricao') or 'Material'}",
                quantidade=quantidade,
                total=subtotal,
                detalhe=item.get("almoxarifado") or "",
            )
            somar_ranking(
                almoxarifados_quantidade,
                item.get("almoxarifado") or "Nao informado",
                quantidade=quantidade,
                total=subtotal,
            )

        recentes.append(
            {
                "id": snapshot.id,
                "glpi": glpi,
                "usuarioNome": nome_usuario,
                "totalEstimado": total,
                "criadoEm": dados.get("criadoEm"),
                "dataLocal": dados.get("dataLocal") or "",
            }
        )

    recentes = sorted(
        recentes,
        key=lambda item: obter_data_ordenacao_solicitacao(item),
        reverse=True,
    )[:5]

    materiais_top = obter_top_ranking(
        materiais_quantidade,
        "quantidade",
        1,
    )

    resumo = {
        "totalSolicitacoes": total_solicitacoes,
        "aguardando": aguardando,
        "vinculadas": vinculadas,
        "totalItens": total_itens,
        "totalEstimado": total_estimado,
        "totalUsuarios": len(usuarios),
        "totalGlpis": len(glpis),
        "totalRequisicoes": len(requisicoes),
        "ticketMedio": (
            total_estimado / total_solicitacoes
            if total_solicitacoes
            else 0
        ),
        "itemMaisSolicitado": (
            materiais_top[0]["label"]
            if materiais_top
            else "Sem item"
        ),
        "usuariosPorSolicitacoes": obter_top_ranking(
            usuarios_quantidade,
            "quantidade",
        ),
        "usuariosPorValor": obter_top_ranking(
            usuarios_valor,
            "total",
        ),
        "glpisPorValor": obter_top_ranking(
            glpis_valor,
            "total",
        ),
        "requisicoesPorValor": obter_top_ranking(
            requisicoes_valor,
            "total",
        ),
        "materiaisPorQuantidade": obter_top_ranking(
            materiais_quantidade,
            "quantidade",
        ),
        "almoxarifadosPorQuantidade": obter_top_ranking(
            almoxarifados_quantidade,
            "quantidade",
        ),
        "recentes": recentes,
        "atualizadoEm": timestamp_servidor(dry_run),
        "origem": "importar_firebase",
    }

    db.collection("resumosAdmin").document("dashboardGeral").set(
        resumo,
        merge=True,
    )

    print("Resumo admin recalculado: resumosAdmin/dashboardGeral")


# =========================
# EXECUCAO
# =========================

def montar_parser():
    carregar_env_local()

    parser = argparse.ArgumentParser(
        description="Importa planilhas do Almoxarifado Inteligente para Firebase."
    )

    parser.add_argument(
        "--service-account",
        default=(
            os.getenv("FIREBASE_SERVICE_ACCOUNT", "") or
            os.getenv("GOOGLE_APPLICATION_CREDENTIALS", "")
        ),
        help="Caminho do JSON de chave do Firebase Admin. Tambem pode usar GOOGLE_APPLICATION_CREDENTIALS.",
    )

    parser.add_argument(
        "--somente",
        nargs="+",
        choices=[
            "usuarios",
            "materiais",
            "requisicoes",
        ],
        default=[
            "usuarios",
            "materiais",
            "requisicoes",
        ],
        help="Define quais importacoes executar.",
    )

    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Le as planilhas e mostra contagens, sem gravar no Firebase.",
    )

    parser.add_argument(
        "--salvar-cpf",
        action="store_true",
        help="Grava CPF no Firestore. Por seguranca, o padrao e nao gravar.",
    )

    parser.add_argument(
        "--atualizar-senhas",
        action="store_true",
        help="Atualiza a senha dos usuarios existentes para o CPF da planilha.",
    )

    parser.add_argument(
        "--admin-matricula",
        nargs="+",
        default=obter_lista_env("ADMIN_MATRICULAS"),
        help="Define uma ou mais matriculas como admin durante a importacao.",
    )

    parser.add_argument(
        "--desativar-materiais-ausentes",
        action="store_true",
        help="Marca como inativos os materiais que existem no Firestore mas nao existem mais na planilha.",
    )

    parser.add_argument(
        "--nao-atualizar-solicitacoes",
        action="store_true",
        help="Importa GLPI x requisicao sem atualizar a colecao solicitacoes.",
    )

    parser.add_argument(
        "--nao-atualizar-resumo-admin",
        action="store_true",
        help="Nao recalcula o documento agregado resumosAdmin/dashboardGeral.",
    )

    return parser


def main():
    parser = montar_parser()
    args = parser.parse_args()

    db = None

    if not args.dry_run:
        inicializar_firebase(args.service_account)
        db = firestore.client()

    print("\n==============================")
    print("IMPORTACAO FIREBASE")
    print("==============================")

    if args.dry_run:
        print("Modo dry-run: nenhuma alteracao sera gravada.")

    if "usuarios" in args.somente:
        print("\nUsuarios")
        importar_usuarios(
            db,
            args.dry_run,
            args.salvar_cpf,
            args.atualizar_senhas,
            args.admin_matricula,
        )

    if "materiais" in args.somente:
        print("\nMateriais")
        importar_materiais(
            db,
            args.dry_run,
            args.desativar_materiais_ausentes,
        )

    if "requisicoes" in args.somente:
        print("\nGLPI x Requisicao")
        mapa = importar_vinculos_requisicoes(
            db,
            args.dry_run,
        )

        if not args.nao_atualizar_solicitacoes:
            atualizar_solicitacoes_com_requisicoes(
                db,
                mapa,
                args.dry_run,
            )

    if not args.nao_atualizar_resumo_admin:
        print("\nResumo admin")
        recalcular_resumo_admin(
            db,
            args.dry_run,
        )

    print("\n==============================")
    print("PROCESSO FINALIZADO")
    print("==============================")


if __name__ == "__main__":
    main()
