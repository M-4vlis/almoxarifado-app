import pandas as pd
import json
import math
from pathlib import Path
import re
import unicodedata
from openpyxl import load_workbook

# =========================
# CONFIGURAÇÕES
# =========================

ARQUIVO_MATERIAIS = "data/LISTA de MATERIAIS.xlsx"

ARQUIVO_COLABORADORES = "data/DADOS COLABORADORES.xlsx"

ARQUIVO_GLPI_REQUISICAO = "data/GLPI - REQUISICAO.xlsx"

ABAS_MATERIAIS = [
    "CENTRAL",
    "SUB",
    "BENFICA"
]

PASTA_DATA = Path("data")

PASTA_DATA.mkdir(exist_ok=True)

# =========================
# FUNÇÕES AUXILIARES
# =========================

def limpar_numero(valor):
    """
    Recebe qualquer valor vindo do Excel e devolve apenas números em texto.
    Usado para códigos de material, GLPI e número de requisição.
    """

    if pd.isna(valor):
        return ""

    texto = str(valor).strip()

    if texto.endswith(".0"):
        texto = texto[:-2]

    texto = re.sub(r"\D", "", texto)

    return texto


def limpar_texto(valor):
    """
    Limpa textos vindos do Excel.
    """

    if pd.isna(valor):
        return ""

    return str(valor).strip()


def normalizar_cabecalho(valor):
    texto = limpar_texto(valor)
    texto = unicodedata.normalize("NFKD", texto)
    texto = "".join(
        caractere
        for caractere in texto
        if not unicodedata.combining(caractere)
    )
    texto = texto.lower()
    texto = re.sub(r"[^a-z0-9]+", " ", texto)
    return texto.strip()


def obter_coluna(colunas, nomes, obrigatoria=True):
    mapa = {
        normalizar_cabecalho(coluna): coluna
        for coluna in colunas
    }

    for nome in nomes:
        coluna = mapa.get(
            normalizar_cabecalho(nome)
        )

        if coluna is not None:
            return coluna

    if obrigatoria:
        raise Exception(
            f"Coluna obrigatoria nao encontrada: {nomes[0]}"
        )

    return None


def converter_valor_brasileiro(valor):
    if pd.isna(valor):
        return 0

    if isinstance(valor, str):
        texto = valor.strip()
        texto = texto.replace("R$", "")
        texto = texto.replace("\xa0", "")
        texto = texto.replace(" ", "")

        if "," in texto and "." in texto:
            if texto.rfind(",") > texto.rfind("."):
                texto = texto.replace(".", "")
                texto = texto.replace(",", ".")
            else:
                texto = texto.replace(",", "")

        elif "," in texto:
            texto = texto.replace(".", "")
            texto = texto.replace(",", ".")

        valor = texto

    try:
        numero = float(valor)
    except (TypeError, ValueError):
        return 0

    if not math.isfinite(numero):
        return 0

    return numero


def formatar_valor_unitario(valor):
    numero = converter_valor_brasileiro(valor)

    texto = f"R$ {numero:,.2f}"
    texto = texto.replace(",", "X")
    texto = texto.replace(".", ",")
    texto = texto.replace("X", ".")

    return texto


def preparar_valor_unitario(valor):
    numero = converter_valor_brasileiro(valor)

    return {
        "valorUnitario": formatar_valor_unitario(numero),
        "valorUnitarioNumero": numero,
    }


def formatar_valor_excel_com_zeros(celula):
    """
    Lê uma célula diretamente do Excel usando openpyxl.

    Objetivo:
    preservar zeros à esquerda quando a planilha usa formatação,
    por exemplo:
    0035
    01234567890

    Isso é especialmente importante para matrícula e CPF.
    """

    valor = celula.value

    if valor is None:
        return ""

    formato = str(celula.number_format)

    # =========================
    # TEXTO NORMAL
    # =========================

    if isinstance(valor, str):

        texto = valor.strip()

        texto = re.sub(
            r"\D",
            "",
            texto
        )

        return texto

    # =========================
    # NÚMERO
    # =========================

    if isinstance(valor, float):

        if valor.is_integer():

            valor = int(valor)

    texto = str(valor).strip()

    if texto.endswith(".0"):
        texto = texto[:-2]

    texto = re.sub(
        r"\D",
        "",
        texto
    )

    # =========================
    # VERIFICA FORMATO COM ZEROS
    # =========================
    # Exemplos:
    # 0000
    # 000.000.000-00
    # 00\.000\.000\/0000-00

    quantidade_zeros_formato =        formato.count("0")

    if quantidade_zeros_formato > len(texto):

        texto =            texto.zfill(
                quantidade_zeros_formato
            )

    return texto


def salvar_json(nome_arquivo, dados):
    """
    Salva uma lista/dicionário em JSON dentro da pasta data.
    """

    caminho =        PASTA_DATA / nome_arquivo

    with open(
        caminho,
        "w",
        encoding="utf-8"
    ) as arquivo:

        json.dump(
            dados,
            arquivo,
            ensure_ascii=False,
            indent=4
        )

    print(
        f"Arquivo salvo em: {caminho}"
    )


# =========================
# GERA MATERIAIS.JSON
# =========================

def gerar_materiais_json():

    print("\n==============================")
    print("GERANDO materiais.json")
    print("==============================")

    materiais_dict = {}

    for aba in ABAS_MATERIAIS:

        print(f"\nLendo aba: {aba}")

        df = pd.read_excel(
            ARQUIVO_MATERIAIS,
            sheet_name=aba
        )

        coluna_codigo = obter_coluna(
            df.columns,
            [
                "Cód",
                "Cod",
                "Codigo",
                "Código",
            ],
        )

        coluna_material = obter_coluna(
            df.columns,
            [
                "Material",
            ],
        )

        coluna_estoque = obter_coluna(
            df.columns,
            [
                "Estoque",
            ],
        )

        coluna_valor_unitario = obter_coluna(
            df.columns,
            [
                "Valor Unitário",
                "Valor unitário",
                "VALOR UNITÁRIO",
                "Valor Unitario",
                "Valor",
                "VALOR",
            ],
            obrigatoria=False,
        )

        if not coluna_valor_unitario:
            print(
                "AVISO: Coluna de valor unitario nao encontrada "
                f"na aba {aba}. Gerando valores como 0."
            )

        df["CÃ³d"] = df[coluna_codigo]
        df["Material"] = df[coluna_material]
        df["Estoque"] = df[coluna_estoque]

        df = df[
            [
                "Cód",
                "Material",
                "Estoque",
                *(
                    [coluna_valor_unitario]
                    if coluna_valor_unitario
                    else []
                )
            ]
        ]

        df = df.dropna(
            subset=[
                "Cód",
                "Material"
            ]
        )

        for _, linha in df.iterrows():

            codigo =                limpar_numero(
                    linha["Cód"]
                )

            descricao =                limpar_texto(
                    linha["Material"]
                )

            estoque =                linha["Estoque"]

            if pd.isna(estoque):

                estoque = 0

            try:

                estoque =                    float(estoque)

            except ValueError:

                estoque = 0

            disponivel =                estoque > 0

            valor_unitario =                preparar_valor_unitario(0)

            if coluna_valor_unitario:

                valor_unitario =                    preparar_valor_unitario(
                        linha[coluna_valor_unitario]
                    )

            chave =                f"{codigo}_{aba}"

            if chave not in materiais_dict:

                materiais_dict[chave] = {

                    "codigo": codigo,

                    "descricao": descricao,

                    "almoxarifado": aba,

                    "disponivel": disponivel,

                    "valorUnitario": valor_unitario["valorUnitario"],

                    "valorUnitarioNumero": valor_unitario["valorUnitarioNumero"]

                }

            else:

                material_existente =                    materiais_dict[chave]

                if disponivel:

                    material_existente["disponivel"] =                        True

                if len(descricao) > len(
                    material_existente["descricao"]
                ):

                    material_existente["descricao"] =                        descricao

                if (
                    valor_unitario["valorUnitarioNumero"] > 0 and
                    converter_valor_brasileiro(
                        material_existente.get("valorUnitarioNumero") or
                        material_existente.get("valorUnitario")
                    ) <= 0
                ):

                    material_existente["valorUnitario"] =                        valor_unitario["valorUnitario"]

                    material_existente["valorUnitarioNumero"] =                        valor_unitario["valorUnitarioNumero"]

    materiais =        list(
            materiais_dict.values()
        )

    salvar_json(
        "materiais.json",
        materiais
    )

    print(
        f"Total de materiais únicos: {len(materiais)}"
    )


# =========================
# GERA COLABORADORES.JSON
# =========================

def gerar_colaboradores_json():

    print("\n==============================")
    print("GERANDO colaboradores.json")
    print("==============================")

    caminho =        Path(
            ARQUIVO_COLABORADORES
        )

    if not caminho.exists():

        print(
            "Arquivo de colaboradores não encontrado. Pulando geração."
        )

        return

    workbook =        load_workbook(
            caminho,
            data_only=True
        )

    planilha =        workbook.active

    cabecalhos = {}

    for coluna in range(
        1,
        planilha.max_column + 1
    ):

        valor =            planilha.cell(
                row=1,
                column=coluna
            ).value

        if valor is not None:

            cabecalhos[
                str(valor).strip()
            ] = coluna

    colunas_necessarias = [
        "NOME",
        "MATRÍCULA",
        "CPF"
    ]

    for coluna in colunas_necessarias:

        if coluna not in cabecalhos:

            raise Exception(
                f"Coluna obrigatória não encontrada na planilha de colaboradores: {coluna}"
            )

    colaboradores = []

    for linha in range(
        2,
        planilha.max_row + 1
    ):

        celula_nome =            planilha.cell(
                row=linha,
                column=cabecalhos["NOME"]
            )

        celula_matricula =            planilha.cell(
                row=linha,
                column=cabecalhos["MATRÍCULA"]
            )

        celula_cpf =            planilha.cell(
                row=linha,
                column=cabecalhos["CPF"]
            )

        nome =            limpar_texto(
                celula_nome.value
            )

        matricula =            formatar_valor_excel_com_zeros(
                celula_matricula
            )

        cpf =            formatar_valor_excel_com_zeros(
                celula_cpf
            )

        if not nome or not matricula or not cpf:

            continue

        colaboradores.append({

            "nome": nome,

            "matricula": matricula,

            "cpf": cpf

        })

    salvar_json(
        "colaboradores.json",
        colaboradores
    )

    print(
        f"Total de colaboradores: {len(colaboradores)}"
    )


# =========================
# GERA REQUISICOES.JSON
# =========================

def gerar_requisicoes_json():

    print("\n==============================")
    print("GERANDO requisicoes.json")
    print("==============================")

    caminho =        Path(
            ARQUIVO_GLPI_REQUISICAO
        )

    if not caminho.exists():

        print(
            "Arquivo GLPI x Requisição não encontrado. Pulando geração."
        )

        return

    df = pd.read_excel(
            ARQUIVO_GLPI_REQUISICAO
        )

    colunas_necessarias = [
        "GLPI",
        "NÚMERO DA REQUISIÇÃO"
    ]

    for coluna in colunas_necessarias:

        if coluna not in df.columns:

            raise Exception(
                f"Coluna obrigatória não encontrada na planilha GLPI x Requisição: {coluna}"
            )

    requisicoes = []

    df = df.dropna(
            subset=[
                "GLPI",
                "NÚMERO DA REQUISIÇÃO"
            ]
        )

    for _, linha in df.iterrows():

        glpi =            limpar_numero(
                linha["GLPI"]
            )

        numero_requisicao =            limpar_numero(
                linha["NÚMERO DA REQUISIÇÃO"]
            )

        if not glpi or not numero_requisicao:

            continue

        requisicoes.append({

            "glpi": glpi,

            "numero_requisicao": numero_requisicao

        })

    salvar_json(
        "requisicoes.json",
        requisicoes
    )

    print(
        f"Total de vínculos GLPI x Requisição: {len(requisicoes)}"
    )


# =========================
# EXECUÇÃO PRINCIPAL
# =========================

def main():

    gerar_materiais_json()

    gerar_colaboradores_json()

    gerar_requisicoes_json()

    print("\n==============================")
    print("PROCESSO FINALIZADO COM SUCESSO")
    print("==============================")


if __name__ == "__main__":

    main()
